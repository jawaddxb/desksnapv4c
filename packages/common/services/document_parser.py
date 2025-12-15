"""
Document Parser Service
Extracts text from various document formats.
KISS: Simple extraction, plain text output.
"""
import csv
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_text(file_path: str, file_type: str) -> str:
    """
    Extract plain text from a document file.

    Args:
        file_path: Path to the document file
        file_type: Type of document (pdf, docx, xlsx, csv, txt, md)

    Returns:
        Extracted text content as a single string
    """
    file_type = file_type.lower()

    try:
        if file_type == "pdf":
            return extract_pdf(file_path)
        elif file_type == "docx":
            return extract_docx(file_path)
        elif file_type == "xlsx":
            return extract_xlsx(file_path)
        elif file_type == "csv":
            return extract_csv(file_path)
        elif file_type in ("txt", "md"):
            return extract_text_file(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
    except Exception as e:
        logger.error(f"Error extracting text from {file_type}: {e}")
        raise


def extract_pdf(file_path: str) -> str:
    """Extract text from PDF using pymupdf."""
    try:
        import fitz  # pymupdf
    except ImportError:
        raise ImportError("pymupdf is required for PDF extraction. Install with: pip install pymupdf")

    text_parts = []
    try:
        doc = fitz.open(file_path)
        for page_num, page in enumerate(doc):
            page_text = page.get_text()
            if page_text.strip():
                text_parts.append(f"[Page {page_num + 1}]\n{page_text}")
        doc.close()
    except Exception as e:
        logger.error(f"Error reading PDF: {e}")
        raise ValueError(f"Failed to read PDF: {str(e)}")

    return "\n\n".join(text_parts)


def extract_docx(file_path: str) -> str:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
    except ImportError:
        raise ImportError("python-docx is required for DOCX extraction. Install with: pip install python-docx")

    text_parts = []
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                text_parts.append(text)
    except Exception as e:
        logger.error(f"Error reading DOCX: {e}")
        raise ValueError(f"Failed to read DOCX: {str(e)}")

    return "\n\n".join(text_parts)


def extract_xlsx(file_path: str) -> str:
    """Extract text from XLSX using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for XLSX extraction. Install with: pip install openpyxl")

    text_parts = []
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_content = [f"[Sheet: {sheet_name}]"]

            # Get header row
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            # Try to detect header row
            header = rows[0] if rows else None
            if header and any(cell is not None for cell in header):
                header_str = " | ".join(str(cell) if cell else "" for cell in header)
                sheet_content.append(f"Headers: {header_str}")

            # Get data rows (limit to prevent huge outputs)
            max_rows = 500
            for row in rows[1:max_rows]:
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):
                    sheet_content.append(" | ".join(row_values))

            if len(rows) > max_rows:
                sheet_content.append(f"... ({len(rows) - max_rows} more rows)")

            text_parts.append("\n".join(sheet_content))

        wb.close()
    except Exception as e:
        logger.error(f"Error reading XLSX: {e}")
        raise ValueError(f"Failed to read XLSX: {str(e)}")

    return "\n\n".join(text_parts)


def extract_csv(file_path: str) -> str:
    """Extract text from CSV."""
    text_parts = []
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            rows = list(reader)

            if not rows:
                return ""

            # Header
            header = rows[0]
            text_parts.append(f"Headers: {' | '.join(header)}")

            # Data rows (limit to prevent huge outputs)
            max_rows = 500
            for row in rows[1:max_rows]:
                if any(cell.strip() for cell in row):
                    text_parts.append(" | ".join(row))

            if len(rows) > max_rows:
                text_parts.append(f"... ({len(rows) - max_rows} more rows)")

    except Exception as e:
        logger.error(f"Error reading CSV: {e}")
        raise ValueError(f"Failed to read CSV: {str(e)}")

    return "\n".join(text_parts)


def extract_text_file(file_path: str) -> str:
    """Extract text from plain text or markdown file."""
    try:
        return Path(file_path).read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        logger.error(f"Error reading text file: {e}")
        raise ValueError(f"Failed to read text file: {str(e)}")


def estimate_tokens(text: str) -> int:
    """
    Estimate token count for text.
    Rough estimate: ~4 characters per token for English text.
    """
    if not text:
        return 0
    return len(text) // 4


def get_file_type(filename: str) -> str | None:
    """Get file type from filename extension."""
    ext = Path(filename).suffix.lower().lstrip(".")
    supported = {"pdf", "docx", "xlsx", "csv", "txt", "md"}
    return ext if ext in supported else None
